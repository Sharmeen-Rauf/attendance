from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from django.conf import settings
from datetime import datetime, date
from .models import Employee, AttendanceRecord
from .serializers import EmployeeSerializer, AttendanceRecordSerializer, AttendanceActionSerializer
import uuid
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from django.http import HttpResponse


@api_view(['GET'])
def get_server_time(request):
    """Get current server time"""
    return Response({
        'serverTime': timezone.now().isoformat()
    })


@api_view(['GET'])
def get_employees(request):
    """Get list of all employees"""
    employees = Employee.objects.all()
    serializer = EmployeeSerializer(employees, many=True)
    return Response(serializer.data)


@api_view(['POST'])
def submit_attendance(request):
    """Submit attendance action (checkin, breakin, breakout, checkout)"""
    serializer = AttendanceActionSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    employee_id = data['employeeId']
    employee_name = data['employeeName']
    action = data['action']
    timestamp = data['timestamp']
    
    # Convert timestamp to timezone-aware datetime
    if isinstance(timestamp, str):
        timestamp = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
    
    if not timezone.is_aware(timestamp):
        timestamp = timezone.make_aware(timestamp)
    
    today = timestamp.date()
    
    # Get or create today's attendance record
    record, created = AttendanceRecord.objects.get_or_create(
        employee_id=employee_id,
        date=today,
        defaults={
            'id': str(uuid.uuid4()),
            'employee_name': employee_name,
        }
    )
    
    # Update record based on action
    if action == 'checkin':
        if record.check_in_time:
            return Response(
                {'error': 'Already checked in today'},
                status=status.HTTP_400_BAD_REQUEST
            )
        record.check_in_time = timestamp
        record.status = record.calculate_status()
    
    elif action == 'checkout':
        if not record.check_in_time:
            return Response(
                {'error': 'Please check in first'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if record.check_out_time:
            return Response(
                {'error': 'Already checked out today'},
                status=status.HTTP_400_BAD_REQUEST
            )
        # Check if break is completed
        if record.break_in_time and not record.break_out_time:
            return Response(
                {'error': 'Please complete break first'},
                status=status.HTTP_400_BAD_REQUEST
            )
        record.check_out_time = timestamp
    
    elif action == 'breakin':
        if not record.check_in_time:
            return Response(
                {'error': 'Please check in first'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if record.break_in_time:
            return Response(
                {'error': 'Break already started'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if record.break_out_time:
            return Response(
                {'error': 'Only one break allowed per day'},
                status=status.HTTP_400_BAD_REQUEST
            )
        record.break_in_time = timestamp
    
    elif action == 'breakout':
        if not record.break_in_time:
            return Response(
                {'error': 'Break not started'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if record.break_out_time:
            return Response(
                {'error': 'Break already completed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        record.break_out_time = timestamp
        # Calculate break duration
        record.break_duration = record.calculate_break_duration()
        
        # Check if break exceeds max duration
        max_break_seconds = settings.MAX_BREAK_DURATION_MINUTES * 60
        if record.break_duration and record.break_duration > max_break_seconds:
            # Still save but will be flagged in reports
            pass
    
    record.save()
    
    serializer = AttendanceRecordSerializer(record)
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['GET'])
def get_attendance(request):
    """Get attendance records with filters"""
    employee_id = request.query_params.get('employee_id')
    filter_date = request.query_params.get('date')
    filter_status = request.query_params.get('status')
    
    records = AttendanceRecord.objects.all()
    
    if employee_id:
        records = records.filter(employee_id=employee_id)
    
    if filter_date:
        try:
            date_obj = datetime.strptime(filter_date, '%Y-%m-%d').date()
            records = records.filter(date=date_obj)
        except ValueError:
            pass
    
    if filter_status:
        records = records.filter(status=filter_status)
    
    # Order by date descending
    records = records.order_by('-date', '-check_in_time')
    
    serializer = AttendanceRecordSerializer(records, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def get_today_status(request, employee_id):
    """Get today's attendance status for an employee"""
    today = date.today()
    
    try:
        record = AttendanceRecord.objects.get(
            employee_id=employee_id,
            date=today
        )
        serializer = AttendanceRecordSerializer(record)
        return Response(serializer.data)
    except AttendanceRecord.DoesNotExist:
        return Response({
            'checkInTime': None,
            'checkOutTime': None,
            'breakInTime': None,
            'breakOutTime': None,
        })


@api_view(['GET'])
def export_attendance(request):
    """Export attendance records to Excel with color coding"""
    employee_id = request.query_params.get('employee_id')
    filter_date = request.query_params.get('date')
    filter_status = request.query_params.get('status')
    
    records = AttendanceRecord.objects.all()
    
    if employee_id:
        records = records.filter(employee_id=employee_id)
    
    if filter_date:
        try:
            date_obj = datetime.strptime(filter_date, '%Y-%m-%d').date()
            records = records.filter(date=date_obj)
        except ValueError:
            pass
    
    if filter_status:
        records = records.filter(status=filter_status)
    
    records = records.order_by('-date', 'employee_name')
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Headers
    headers = ['Employee Name', 'Employee ID', 'Date', 'Check-In', 'Check-Out', 
               'Break Duration', 'Status']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for record in records:
        check_in = record.check_in_time.strftime('%H:%M:%S') if record.check_in_time else '-'
        check_out = record.check_out_time.strftime('%H:%M:%S') if record.check_out_time else '-'
        
        break_duration = '-'
        if record.break_duration:
            minutes = record.break_duration // 60
            seconds = record.break_duration % 60
            break_duration = f"{minutes}m {seconds}s"
        
        status_text = dict(AttendanceRecord.STATUS_CHOICES).get(record.status, record.status)
        
        row = [
            record.employee_name,
            record.employee_id,
            record.date.strftime('%Y-%m-%d'),
            check_in,
            check_out,
            break_duration,
            status_text,
        ]
        ws.append(row)
        
        # Color code status
        status_cell = ws.cell(row=ws.max_row, column=7)
        if record.status == 'on_time':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100")
        elif record.status == 'grace':
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(color="9C6500")
        elif record.status == 'late':
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006")
        
        # Highlight long breaks
        if record.break_duration and record.break_duration > settings.MAX_BREAK_DURATION_MINUTES * 60:
            break_cell = ws.cell(row=ws.max_row, column=6)
            break_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            break_cell.font = Font(color="9C0006", bold=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'attendance_{filter_date or "all"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

